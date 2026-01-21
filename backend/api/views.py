from rest_framework import viewsets, status
from rest_framework.decorators import api_view, action, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from django.db.models import Count, Sum, Q, Prefetch
from django.http import HttpResponse
from django.utils import timezone
from decimal import Decimal
import csv
import io
import pandas as pd

from .models import Aggregator, Category, Product, Price, Recommendation, PriceHistory, ProductLink, ImportJob, City


class StandardResultsSetPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 1000


from .serializers import (
    AggregatorSerializer,
    CitySerializer,
    CategorySerializer,
    CategoryTreeSerializer,
    ProductSerializer,
    ProductComparisonSerializer,
    ProductLinkSerializer,
    RecommendationSerializer,
    PriceHistorySerializer,
    ImportJobSerializer,
    DashboardSerializer,
)
from .services.importer import DataImporter
from .services.matching import ProductMatcher


class AggregatorViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Aggregator.objects.all()
    serializer_class = AggregatorSerializer


class CityViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = City.objects.all()
    serializer_class = CitySerializer


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.annotate(product_count=Count('product')).order_by('sort_order', 'name')
    serializer_class = CategorySerializer

    @action(detail=False, methods=['get'])
    def tree(self, request):
        """Получить иерархическое дерево категорий"""
        root_categories = Category.objects.filter(parent__isnull=True).order_by('sort_order', 'name')
        serializer = CategoryTreeSerializer(root_categories, many=True)
        return Response(serializer.data)


class ProductViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Product.objects.all().select_related('category')
    serializer_class = ProductSerializer

    @action(detail=False, methods=['get'])
    def comparison(self, request):
        city_slug = request.GET.get('city')
        category_ids = request.GET.getlist('category_ids[]')
        
        price_qs = Price.objects.select_related('aggregator')
        if city_slug:
            price_qs = price_qs.filter(city__slug=city_slug)

        products = Product.objects.all().select_related('category')

        if category_ids:
            try:
                all_ids = set(map(int, category_ids))
                current_batch = list(all_ids)
                
                while current_batch:
                    # Get next level of children
                    children_ids = list(Category.objects.filter(parent_id__in=current_batch).values_list('id', flat=True))
                    # Filter out those we already have
                    new_ids = [cid for cid in children_ids if cid not in all_ids]
                    all_ids.update(new_ids)
                    current_batch = new_ids
                
                products = products.filter(category_id__in=all_ids)
            except Exception as e:
                print(f"Error filtering categories: {e}")

        # Prefetch prices and links
        products = products.prefetch_related(
            Prefetch('price_set', queryset=price_qs, to_attr='filtered_prices'),
            'links'
        )

        # Apply sorting: 1. Number of prices (matches) desc, 2. Priority? 
        # For now, let's just sort by number of prices found in prefetch
        # Actually, let's do matching density sorting in Python or with Annotation
        products = products.annotate(
            price_count=Count('price', filter=Q(price__is_available=True) & (Q(price__city__slug=city_slug) if city_slug else Q()))
        ).order_by('-price_count', 'name')

        # Apply pagination
        paginator = StandardResultsSetPagination()
        page = paginator.paginate_queryset(products, request)

        # Get list of all aggregators for the frontend to show all columns
        aggregators = AggregatorSerializer(Aggregator.objects.all(), many=True).data

        if page is not None:
            serializer = ProductComparisonSerializer(page, many=True, context={'request': request})
            response_data = paginator.get_paginated_response(serializer.data)
            response_data.data['meta'] = {
                'aggregators': aggregators
            }
            return response_data

        serializer = ProductComparisonSerializer(products, many=True, context={'request': request})
        return Response({
            'results': serializer.data,
            'meta': {
                'aggregators': aggregators
            }
        })


class RecommendationViewSet(viewsets.ModelViewSet):
    queryset = Recommendation.objects.all().select_related('product', 'product__category')
    serializer_class = RecommendationSerializer

    @action(detail=True, methods=['post'])
    def apply(self, request, pk=None):
        recommendation = self.get_object()

        if recommendation.status == 'APPLIED':
            return Response(
                {'error': 'Recommendation already applied'},
                status=status.HTTP_400_BAD_REQUEST
            )

        our_aggregator = Aggregator.objects.filter(is_our_company=True).first()

        if recommendation.action_type == 'LOWER_PRICE':
            price_obj = Price.objects.filter(
                product=recommendation.product,
                aggregator=our_aggregator,
                city=recommendation.city
            ).first()

            if price_obj:
                old_price = price_obj.price
                price_obj.price = recommendation.recommended_price
                price_obj.save()

                PriceHistory.objects.create(
                    product=recommendation.product,
                    aggregator=our_aggregator,
                    old_price=old_price,
                    new_price=recommendation.recommended_price
                )

        elif recommendation.action_type == 'ADD_PRODUCT':
            Price.objects.update_or_create(
                product=recommendation.product,
                aggregator=our_aggregator,
                city=recommendation.city,
                defaults={
                    'price': recommendation.recommended_price,
                    'is_available': True
                }
            )

        recommendation.status = 'APPLIED'
        recommendation.save()

        return Response({'status': 'success', 'message': 'Recommendation applied'})

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        recommendation = self.get_object()
        recommendation.status = 'REJECTED'
        recommendation.save()
        return Response({'status': 'success', 'message': 'Recommendation rejected'})


@api_view(['GET'])
def dashboard_stats(request):
    """Optimized dashboard stats using prefetch instead of N+1 queries"""
    city_slug = request.GET.get('city')

    # Build city filter for prices
    price_filter = Q(is_available=True, price__isnull=False)
    if city_slug:
        price_filter &= Q(city__slug=city_slug)

    # Prefetch all prices in one query
    price_prefetch = Prefetch(
        'price_set',
        queryset=Price.objects.filter(price_filter).select_related('aggregator'),
        to_attr='cached_prices'
    )

    products = Product.objects.prefetch_related(price_prefetch)

    total_products = 0
    products_at_top = 0
    products_need_action = 0
    missing_products = 0

    for product in products:
        total_products += 1
        prices = product.cached_prices

        our_price = None
        competitor_prices = []

        for price in prices:
            if price.aggregator.is_our_company:
                our_price = float(price.price)
            else:
                competitor_prices.append(float(price.price))

        if our_price is None:
            missing_products += 1
        elif competitor_prices:
            min_competitor = min(competitor_prices)
            if our_price < min_competitor:
                products_at_top += 1
            else:
                products_need_action += 1
        else:
            # We have product but no competitors - count as top
            products_at_top += 1

    recommendation_qs = Recommendation.objects.filter(status='PENDING')
    if city_slug:
        recommendation_qs = recommendation_qs.filter(city__slug=city_slug)

    pending_recommendations = recommendation_qs.count()
    potential_savings = recommendation_qs.filter(
        potential_savings__isnull=False
    ).aggregate(total=Sum('potential_savings'))['total'] or Decimal('0')

    total_with_our_price = total_products - missing_products
    market_coverage = (total_with_our_price / total_products * 100) if total_products > 0 else 0
    price_competitiveness = (products_at_top / total_with_our_price * 100) if total_with_our_price > 0 else 0

    # Calculate aggregator stats
    aggregator_counts = {}
    aggregators = Aggregator.objects.all()
    for agg in aggregators:
        if not agg.is_our_company:
            # Count products that have an available price from this aggregator
            count = Price.objects.filter(price_filter, aggregator=agg).values('product').distinct().count()
            percent = round((count / total_products * 100), 1) if total_products > 0 else 0
            aggregator_counts[agg.name] = {
                'count': count,
                'percent': percent
            }

    data = {
        'total_products': total_products,
        'products_at_top': products_at_top,
        'products_need_action': products_need_action,
        'missing_products': missing_products,
        'pending_recommendations': pending_recommendations,
        'potential_savings': potential_savings,
        'market_coverage': round(market_coverage, 1),
        'price_competitiveness': round(price_competitiveness, 1),
        'aggregator_stats': aggregator_counts
    }

    return Response(data)


@api_view(['GET'])
def analytics_gaps(request):
    """Get products that we don't have but competitors do - optimized with prefetch"""
    city_slug = request.GET.get('city')

    # Build city filter for prices
    price_filter = Q(is_available=True, price__isnull=False)
    if city_slug:
        price_filter &= Q(city__slug=city_slug)

    # Prefetch all prices in one query
    price_prefetch = Prefetch(
        'price_set',
        queryset=Price.objects.filter(price_filter).select_related('aggregator'),
        to_attr='cached_prices'
    )

    products = Product.objects.select_related('category').prefetch_related(price_prefetch)

    gaps = []
    for product in products:
        prices = product.cached_prices

        our_price = None
        competitor_prices = []

        for price in prices:
            if price.aggregator.is_our_company:
                our_price = price
            else:
                competitor_prices.append(float(price.price))

        # If we don't have this product but competitors do
        if (not our_price or not our_price.price) and competitor_prices:
            min_price = min(competitor_prices)
            gaps.append({
                'product_id': product.id,
                'product_name': product.name,
                'category': product.category.name if product.category else None,
                'min_competitor_price': float(min_price),
                'suggested_price': float(round(min_price - 1, 2))
            })

    return Response(gaps[:100])  # Limit to 100 results for performance


@api_view(['POST'])
def run_algorithm(request):
    """Run the pricing optimization algorithm - optimized with prefetch"""
    city_slug = request.GET.get('city')

    # Build city filter for prices
    price_filter = Q(is_available=True, price__isnull=False)
    if city_slug:
        price_filter &= Q(city__slug=city_slug)

    # Prefetch all prices in one query
    price_prefetch = Prefetch(
        'price_set',
        queryset=Price.objects.filter(price_filter).select_related('aggregator'),
        to_attr='cached_prices'
    )

    products = Product.objects.prefetch_related(price_prefetch)

    matcher = ProductMatcher()
    new_recommendations = []

    for product in products:
        # Pass cached prices to matcher
        rec = matcher.run_with_cached_prices(product, product.cached_prices, city_slug)
        if rec:
            new_recommendations.append(RecommendationSerializer(rec).data)

    return Response({
        'status': 'success',
        'new_recommendations': len(new_recommendations),
        'recommendations': new_recommendations[:50]  # Limit response size
    })


@api_view(['GET'])
def price_history(request):
    history = PriceHistory.objects.all().select_related('product', 'aggregator').order_by('-changed_at')[:50]
    serializer = PriceHistorySerializer(history, many=True)
    return Response(serializer.data)


class ProductLinkViewSet(viewsets.ModelViewSet):
    queryset = ProductLink.objects.all().select_related('product', 'aggregator')
    serializer_class = ProductLinkSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        product_id = self.request.query_params.get('product_id')
        if product_id:
            queryset = queryset.filter(product_id=product_id)
        return queryset


class ImportJobViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ImportJob.objects.all().order_by('-created_at')
    serializer_class = ImportJobSerializer


@api_view(['GET'])
def import_template(request, template_type):
    """Скачать Excel шаблон для импорта"""
    # Define columns
    templates = {
        'products': ['name', 'category', 'brand', 'manufacturer', 'country_of_origin', 'weight_value', 'weight_unit', 'sku', 'image_url'],
        'prices': ['product_name_or_sku', 'aggregator', 'price', 'is_available', 'competitor_brand', 'competitor_country'],
        'links': ['product_name_or_sku', 'aggregator', 'url', 'external_name'],
        'categories': ['name', 'parent_name', 'icon', 'sort_order'],
    }

    if template_type not in templates:
        return Response({'error': 'Unknown template type'}, status=400)

    # create DataFrame
    df = pd.DataFrame(columns=templates[template_type])

    # Add examples
    examples = {
        'products': [{
            'name': 'Молоко 2.5%', 
            'category': 'Молочные продукты', 
            'brand': 'Lactel', 
            'manufacturer': 'Lactalis', 
            'country_of_origin': 'Казахстан', 
            'weight_value': 1, 
            'weight_unit': 'l', 
            'sku': 'MLK001', 
            'image_url': ''
        }],
        'prices': [{
            'product_name_or_sku': 'Молоко 2.5%', 
            'aggregator': 'Glovo', 
            'price': 450, 
            'is_available': True,
            'competitor_brand': 'Lactel', # Optional: if different
            'competitor_country': 'Германия' # Optional
        }],
        'links': [{
            'product_name_or_sku': 'Молоко 2.5%', 
            'aggregator': 'Glovo', 
            'url': 'https://glovo.kz/store/product/123', 
            'external_name': 'Молоко Lactel 2.5% 1л'
        }],
        'categories': [{
            'name': 'Йогурты', 
            'parent_name': 'Молочные продукты', 
            'icon': 'yogurt', 
            'sort_order': 1
        }],
    }
    
    if template_type in examples:
        df = pd.concat([df, pd.DataFrame(examples[template_type])], ignore_index=True)

    # Write to Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Template')
        
        # Adjust column widths for better UX
        worksheet = writer.sheets['Template']
        for idx, col in enumerate(df.columns):
            max_len = max(
                df[col].astype(str).map(len).max(),
                len(str(col))
            ) + 2
            worksheet.column_dimensions[chr(65 + idx)].width = min(max_len, 50)

    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{template_type}_template.xlsx"'
    return response


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def import_products(request):
    """Импорт товаров из CSV/Excel"""
    file = request.FILES.get('file')
    if not file:
        return Response({'error': 'No file provided'}, status=400)

    job = ImportJob.objects.create(
        job_type='products',
        file_name=file.name,
        status='processing'
    )

    importer = DataImporter(job)
    # Run in background ideally, but synchronous for now as in original
    importer.process(file)

    return Response({
        'job_id': job.id,
        'status': job.status,
        'total': job.total_rows,
        'success': job.success_count,
        'errors': job.error_count,
        'error_details': job.error_details
    })


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def import_prices(request):
    """Импорт цен из CSV/Excel"""
    file = request.FILES.get('file')
    if not file:
        return Response({'error': 'No file provided'}, status=400)

    job = ImportJob.objects.create(
        job_type='prices',
        file_name=file.name,
        status='processing'
    )

    importer = DataImporter(job)
    importer.process(file)

    return Response({
        'job_id': job.id,
        'status': job.status,
        'total': job.total_rows,
        'success': job.success_count,
        'errors': job.error_count,
        'error_details': job.error_details
    })


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def import_links(request):
    """Импорт ссылок из CSV/Excel"""
    file = request.FILES.get('file')
    if not file:
        return Response({'error': 'No file provided'}, status=400)

    job = ImportJob.objects.create(
        job_type='links',
        file_name=file.name,
        status='processing'
    )

    importer = DataImporter(job)
    importer.process(file)

    return Response({
        'job_id': job.id,
        'status': job.status,
        'total': job.total_rows,
        'success': job.success_count,
        'errors': job.error_count,
        'error_details': job.error_details
    })


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def import_categories(request):
    """Импорт категорий из CSV/Excel"""
    file = request.FILES.get('file')
    if not file:
        return Response({'error': 'No file provided'}, status=400)

    job = ImportJob.objects.create(
        job_type='categories',
        file_name=file.name,
        status='processing'
    )

    importer = DataImporter(job)
    importer.process(file)

    return Response({
        'job_id': job.id,
        'status': job.status,
        'total': job.total_rows,
        'success': job.success_count,
        'errors': job.error_count,
        'error_details': job.error_details
    })


@api_view(['GET'])
def json_import_info(request):
    """Get info about available JSON files for import"""
    from .services.json_importer import JSONDataImporter, CATEGORY_PATTERNS
    
    importer = JSONDataImporter()
    files = importer.list_available_files()
    
    categories = []
    for slug, info in CATEGORY_PATTERNS.items():
        categories.append({
            'slug': slug,
            'name': info['name'],
            'subcategories': list(info['subcategories'].keys())
        })
    
    return Response({
        'files': files,
        'categories': categories,
    })


@api_view(['POST'])
def import_from_json(request):
    """Import products from JSON files in Data folder"""
    from .services.json_importer import JSONDataImporter
    
    data = request.data
    categories = data.get('categories', None)
    aggregators = data.get('aggregators', None)
    limit = data.get('limit', None)
    dry_run = data.get('dry_run', False)
    
    if limit:
        try:
            limit = int(limit)
        except:
            limit = None
    
    job = ImportJob.objects.create(
        job_type='products',
        file_name='JSON Data Import',
        status='processing'
    )
    
    importer = JSONDataImporter(job)
    result = importer.import_from_json(
        categories=categories,
        aggregators=aggregators,
        limit_per_category=limit,
        dry_run=dry_run,
    )
    
    return Response({
        'job_id': job.id,
        'status': result['status'],
        'total_imported': result['total_imported'],
        'categories': result['categories'],
        'aggregators': result['aggregators'],
        'stats': result.get('stats', {}),
        'errors': result['errors'][:10] if result['errors'] else [],
    })


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def upload_custom_json(request):
    """Upload and import from custom JSON files"""
    from .services.json_importer import JSONDataImporter
    import json
    import tempfile
    import os
    
    files = request.FILES.getlist('files')
    if not files:
        return Response({'error': 'No files provided'}, status=400)
    
    # Parse form data
    categories = request.data.get('categories', '[]')
    if isinstance(categories, str):
        try:
            categories = json.loads(categories)
        except:
            categories = []
    
    limit = request.data.get('limit', 100)
    try:
        limit = int(limit)
    except:
        limit = 100
    
    dry_run = request.data.get('dry_run', 'false')
    dry_run = dry_run == 'true' or dry_run == True
    
    # Create import job
    job = ImportJob.objects.create(
        job_type='products',
        file_name=f'Custom JSON Import ({len(files)} files)',
        status='processing'
    )
    
    importer = JSONDataImporter(job)
    
    # Save uploaded files to temp directory and process
    temp_dir = tempfile.mkdtemp()
    aggregator_results = {}
    total_imported = 0
    all_errors = []
    
    try:
        for uploaded_file in files:
            # Save to temp file
            file_path = os.path.join(temp_dir, uploaded_file.name)
            with open(file_path, 'wb') as f:
                for chunk in uploaded_file.chunks():
                    f.write(chunk)
            
            # Determine aggregator name from filename
            agg_name = uploaded_file.name.replace('.json', '').replace('_products', '').replace('.', '_')
            
            # Process the file
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                if not isinstance(data, list):
                    all_errors.append(f'{uploaded_file.name}: Invalid format - expected array')
                    continue
                
                file_count = 0
                for item in data:
                    if limit and file_count >= limit:
                        break
                    
                    # Parse product
                    parsed = importer.parse_product(item, agg_name)
                    if not parsed:
                        continue
                    
                    # Check category if specified
                    if categories:
                        category_info = importer.detect_category(item)
                        if not category_info or category_info.get('slug') not in categories:
                            # Also check if custom category matches
                            title_lower = parsed.get('title', '').lower()
                            matched = False
                            for custom_cat in categories:
                                if custom_cat.lower() in title_lower:
                                    matched = True
                                    category_info = {'name': custom_cat, 'slug': custom_cat.lower(), 'subcategory': None}
                                    break
                            if not matched:
                                continue
                    else:
                        category_info = importer.detect_category(item)
                    
                    if not dry_run and category_info:
                        # Create/update product
                        importer.stats['matched_category'] += 1
                        
                        # Get or create category
                        category, _ = Category.objects.get_or_create(
                            name=category_info['name'],
                            defaults={'slug': category_info.get('slug', '')}
                        )
                        
                        # Create product
                        product, created = Product.objects.update_or_create(
                            name=parsed['title'][:500],
                            defaults={
                                'category': category,
                                'brand': parsed.get('brand', ''),
                                'weight_value': parsed.get('weight_value'),
                                'weight_unit': parsed.get('weight_unit', ''),
                                'image_url': parsed.get('image_url', ''),
                            }
                        )
                        
                        # Create price
                        aggregator = importer.ensure_aggregator(agg_name)
                        city = importer.ensure_city(parsed.get('city', 'almaty'))
                        
                        Price.objects.update_or_create(
                            product=product,
                            aggregator=aggregator,
                            city=city,
                            defaults={
                                'price': parsed.get('price'),
                                'is_available': True
                            }
                        )
                        
                        file_count += 1
                        total_imported += 1
                    elif category_info:
                        file_count += 1
                        total_imported += 1
                
                aggregator_results[agg_name] = {'count': file_count}
                
            except Exception as e:
                all_errors.append(f'{uploaded_file.name}: {str(e)}')
    
    finally:
        # Cleanup temp files
        import shutil
        shutil.rmtree(temp_dir, ignore_errors=True)
    
    job.status = 'completed' if not all_errors else 'completed_with_errors'
    job.success_count = total_imported
    job.error_count = len(all_errors)
    job.save()
    
    return Response({
        'job_id': job.id,
        'status': job.status,
        'total_imported': total_imported,
        'categories': {},
        'aggregators': aggregator_results,
        'stats': {'total_read': total_imported},
        'errors': all_errors[:10],
    })


@api_view(['GET'])
def export_products(request):
    """Export comparison table to Excel with formatting"""
    city_slug = request.GET.get('city')
    category_ids = request.GET.getlist('category_ids[]')
    
    products = Product.objects.all().select_related('category')
    
    if category_ids:
        try:
            all_ids = set(map(int, category_ids))
            current_batch = list(all_ids)
            while current_batch:
                children_ids = list(Category.objects.filter(parent_id__in=current_batch).values_list('id', flat=True))
                new_ids = [cid for cid in children_ids if cid not in all_ids]
                all_ids.update(new_ids)
                current_batch = new_ids
            products = products.filter(category_id__in=all_ids)
        except:
            pass
    
    # Get aggregators (competitors only for columns)
    aggregators = list(Aggregator.objects.all().order_by('-is_our_company', 'name'))
    our_agg = next((a for a in aggregators if a.is_our_company), None)
    competitor_aggs = [a for a in aggregators if not a.is_our_company]
    
    price_filter = Q(is_available=True)
    if city_slug:
        price_filter &= Q(city__slug=city_slug)
    
    # Build price filter for annotate (needs price__ prefix)
    annotate_filter = Q(price__is_available=True)
    if city_slug:
        annotate_filter &= Q(price__city__slug=city_slug)
    
    products = products.prefetch_related(
        Prefetch('price_set', queryset=Price.objects.filter(price_filter).select_related('aggregator'))
    ).annotate(
        price_count=Count('price', filter=annotate_filter)
    ).order_by('-price_count', 'name')
    
    # Create Excel with openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Сравнение цен"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    our_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    top_fill = PatternFill(start_color="BBF7D0", end_color="BBF7D0", fill_type="solid")  # Best price
    high_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # Higher than others
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    # Header row
    headers = ['№', 'Товар', 'Категория']
    
    # Our company column first
    if our_agg:
        headers.append(f'🏠 {our_agg.name}')
    
    # Competitor columns
    for agg in competitor_aggs:
        headers.append(agg.name)
    
    headers.extend(['Позиция', 'Мин. цена конк.', 'Рек. цена', 'Статус'])
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Data rows
    row_num = 2
    for idx, product in enumerate(products[:2000], 1):
        # Build price map
        price_map = {}
        if hasattr(product, 'price_set'):
            for price in product.price_set.all():
                price_map[price.aggregator_id] = float(price.price) if price.price else None
        
        # Calculate stats
        our_price = price_map.get(our_agg.id) if our_agg else None
        competitor_prices = [price_map.get(a.id) for a in competitor_aggs if price_map.get(a.id) is not None]
        min_competitor = min(competitor_prices) if competitor_prices else None
        
        # Position and status
        if our_price is None:
            position = '—'
            status = 'Нет товара'
        elif not competitor_prices:
            position = '1'
            status = '✓ Лидер'
        elif our_price < min_competitor:
            position = '1'
            status = '✓ Лидер'
        elif our_price == min_competitor:
            position = '2'
            status = '≈ Равная'
        else:
            all_prices = sorted(set(competitor_prices + [our_price]))
            position = str(all_prices.index(our_price) + 1)
            status = '↑ Выше рынка'
        
        recommended = min_competitor - 1 if min_competitor else None
        
        # Write row
        col = 1
        
        # №
        ws.cell(row=row_num, column=col, value=idx).border = thin_border
        col += 1
        
        # Product name
        cell = ws.cell(row=row_num, column=col, value=product.name[:80])
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True)
        col += 1
        
        # Category
        cell = ws.cell(row=row_num, column=col, value=product.category.name if product.category else '')
        cell.border = thin_border
        col += 1
        
        # Our price
        if our_agg:
            cell = ws.cell(row=row_num, column=col, value=our_price if our_price else '')
            cell.border = thin_border
            cell.fill = our_fill
            cell.alignment = Alignment(horizontal="center")
            if our_price and min_competitor:
                if our_price < min_competitor:
                    cell.fill = top_fill
                elif our_price > min_competitor:
                    cell.fill = high_fill
            col += 1
        
        # Competitor prices
        for agg in competitor_aggs:
            price = price_map.get(agg.id)
            cell = ws.cell(row=row_num, column=col, value=price if price else '')
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            # Highlight min price
            if price and min_competitor and price == min_competitor:
                cell.font = Font(bold=True)
            col += 1
        
        # Position
        cell = ws.cell(row=row_num, column=col, value=position)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        col += 1
        
        # Min competitor price
        cell = ws.cell(row=row_num, column=col, value=min_competitor if min_competitor else '')
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        col += 1
        
        # Recommended price
        cell = ws.cell(row=row_num, column=col, value=recommended if recommended else '')
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        col += 1
        
        # Status
        cell = ws.cell(row=row_num, column=col, value=status)
        cell.border = thin_border
        if '✓' in status:
            cell.font = Font(color="059669")
        elif '↑' in status:
            cell.font = Font(color="DC2626")
        col += 1
        
        row_num += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 5   # №
    ws.column_dimensions['B'].width = 45  # Товар
    ws.column_dimensions['C'].width = 18  # Категория
    
    # Price columns
    for idx in range(4, 4 + len(aggregators)):
        ws.column_dimensions[get_column_letter(idx)].width = 14
    
    # Stats columns
    for idx in range(4 + len(aggregators), 4 + len(aggregators) + 4):
        ws.column_dimensions[get_column_letter(idx)].width = 14
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Write to response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="comparison_export.xlsx"'
    return response


@api_view(['POST'])
def reset_categories(request):
    """Reset and create fresh category structure"""
    from .services.json_importer import JSONDataImporter
    
    # First, nullify all product category references to avoid foreign key constraint
    Product.objects.all().update(category=None)
    
    # Now safely delete all categories
    Category.objects.all().delete()
    
    # Create fresh category structure
    importer = JSONDataImporter()
    categories = importer.ensure_categories()
    
    return Response({
        'status': 'success',
        'categories_created': len(categories),
        'categories': list(categories.keys())
    })


